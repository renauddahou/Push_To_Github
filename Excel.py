from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
#wb = Workbook()
#ws = wb.active
#ws.title = "Data"
#ws.append(['Python', 'Is','Great'])
#wb.save('C:/Users/cathe/OneDrive/Documents/OpenXL.xlsx')
#when saving or loading a workbook, use the forward slash('/') instead of the back slash('\')
wb_2 = load_workbook('C:/Users/cathe/OneDrive/Documents/OpenXL.xlsx')
ws = wb_2['Test']
#wb_2.create_sheet("Test")
#ws.append(['Today', 'Is', 'Friday', 'September', 'tenth', 'Twenty Twenty One'])
#for row in range(1,11):
    #for col in range(1,5):
        #char = get_column_letter(col)
        #print(ws[char + str(row)].value)
#ws.merge_cells('A1:D1')
#ws.unmerge_cells('A1:D1')
#ws.insert_rows(7)
#ws.insert_cols(7)
#ws.delete_cols(7)
#ws.delete_rows(7)
ws.move_range('C1:D11', rows= 2, cols= 2)
wb_2.save('C:/Users/cathe/OneDrive/Documents/OpenXL2.xlsx')